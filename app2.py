from flask import Flask, request, jsonify, render_template
from openpyxl import Workbook, load_workbook
from joblib import load
import pandas as pd
import os

app = Flask(__name__)

# Crear la carpeta de dades
SAVE_FOLDER = 'BITS-X-FIBROSI'
if not os.path.exists(SAVE_FOLDER):
    os.makedirs(SAVE_FOLDER)

# Carregar el model entrenat
model = load('model_death.pkl')
print("Model carregat correctament.")

# Definir les variables del model
model_variables = [
    'Pedigree', 'sex', 'Age at diagnosis', 'FinalDiagnosis', 'TobaccoHistory', 
    'RadiologicalPattern', 'Biopsy', 'Extrapulmonary', 'LungCancer', 'OtherCancer', 
    'NeoplasiaType', 'HematologicAbnormalities', 'BloodCountAbnormalities', 
    'HematologicDiseaseConfirm', 'HematologicDiseaseTypes', 'LiverAbnormalityBefore', 
    'LiverAbnormality', 'LiverDisease', 'LDH', 'ALT', 'AST', 'ALP', 'GGT', 
    'Transaminitis', 'Cholestasis', 'FVC', 'DLCO', 'FirstDegreeRelative', 
    'SecondDegreeRelative', 'MoreThanOneRelative', 'GeneticMutation', 'MutationType', 
    'TelomereShorteningSeverity'
]

# Ruta per a la pàgina principal
@app.route('/')
def index():
    return render_template('app_prova.html')  # Assegura't que aquest fitxer HTML està a /templates

# Ruta per guardar les dades al fitxer Excel
@app.route('/save-excel', methods=['POST'])
def save_excel():
    try:
        data = request.get_json()  # Rebre dades en format JSON
        print(f"Dades rebudes: {data}")  # Per depuració

        if not data:
            return jsonify({'error': 'No hi ha dades per guardar.'}), 400
        
        # Codificar les dades de l'usuari
        tobacco_mapping = {
            'No history of smoking': 0,
            'Active smoker': 1,
            'Ex-smoker': 2
        }

        biopsy_mapping = {
            'biopsy-none': 0,
            'biopsy-endoscopic': 1,
            'biopsy-surgical': 2
        }

        radiological_pattern_mapping = {
            'UIP': 'UIP',
            'Probable UIP': 'Probable UIP',
            'Non UIP': 'Non UIP',
            'Indeterminate UIP': 'Indeterminate UIP'
        }

        extrapulmonary_mapping = {
            'no': 0,
            'yes': 1
        }

        lung_cancer_mapping = {
            'no': 0,
            'yes': 1
        }

        other_cancer_mapping = {
            'no': 0,
            'yes': 1
        }

        hematologic_abnormalities_mapping = {
            'no': 0,
            'yes': 1
        }

        hematologic_disease_confirm_mapping = {
            'no': 0,
            'yes': 1
        }

        liver_abnormality_before_mapping = {
            'no': 0,
            'yes': 1
        }

        liver_abnormality_mapping = {
            'no': 0,
            'yes': 1
        }

        liver_disease_mapping = {
            'no': 0,
            'yes': 1
        }

        genetic_mutation_mapping = {
            'no': 0,
            'yes': 1
        }

        data['TobaccoHistory'] = tobacco_mapping.get(data.get('TobaccoHistory'), -1)
        data['Biopsy'] = biopsy_mapping.get(data.get('Biopsy'), -1)
        data['RadiologicalPattern'] = radiological_pattern_mapping.get(data.get('RadiologicalPattern'), -1)
        data['Extrapulmonary'] = extrapulmonary_mapping.get(data.get('Extrapulmonary'), -1)
        data['LungCancer'] = lung_cancer_mapping.get(data.get('LungCancer'), -1)
        data['OtherCancer'] = other_cancer_mapping.get(data.get('OtherCancer'), -1)
        data['HematologicAbnormalities'] = hematologic_abnormalities_mapping.get(data.get('HematologicAbnormalities'), -1)
        data['HematologicDiseaseConfirm'] = hematologic_disease_confirm_mapping.get(data.get('HematologicDiseaseConfirm'), -1)
        data['LiverAbnormalityBefore'] = liver_abnormality_before_mapping.get(data.get('LiverAbnormalityBefore'), -1)
        data['LiverAbnormality'] = liver_abnormality_mapping.get(data.get('LiverAbnormality'), -1)
        data['LiverDisease'] = liver_disease_mapping.get(data.get('LiverDisease'), -1)
        data['GeneticMutation'] = genetic_mutation_mapping.get(data.get('GeneticMutation'), -1)

        # Crear o carregar el fitxer Excel
        filepath = os.path.join(SAVE_FOLDER, 'respostes_questionari.xlsx')
        if os.path.exists(filepath):
            workbook = load_workbook(filepath)
        else:
            workbook = Workbook()

        # Obtenir o crear la fulla de respostes
        sheet_name = 'Respostes'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        # Afegir els encapçalaments si no existeixen
        if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
            sheet.append(model_variables + ['Predicció'])

        # Afegir la nova fila amb les dades i la predicció
        df_input = pd.DataFrame([data], columns=model_variables)
        df_input = df_input[model_variables]

        # Preprocessar les dades noves per predicció
        df_input = pd.get_dummies(df_input, drop_first=True)

        # Preprocessar les dades noves per predicció
        prediction = model.predict(df_input)[0]
        print(f"Predicció: {prediction}")  # Per depuració

        # Afegir la nova fila amb les dades i la predicció
        user_input = [data.get(var) for var in model_variables] + [prediction]
        sheet.append(user_input)

        # Guardar l'Excel
        workbook.save(filepath)

        return jsonify({'message': 'Dades guardades correctament i predicció feta.', 'predicció': int(prediction)}), 200

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': 'Hi ha hagut un problema en desar les dades o fer la predicció.'}), 500


if __name__ == '__main__':
    app.run(debug=True)