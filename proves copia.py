from flask import Flask, request, jsonify, render_template
from openpyxl import Workbook, load_workbook
from joblib import load
import pandas as pd
import os

app = Flask(__name__)

# Crear la carpeta de dades
SAVE_FOLDER = 'BITS-X-FIBROSI'
if not os.path.exists(SAVE_FOLDER):
    os.makedirs(SAVE_FOLDER)

# Carregar el model entrenat
model_death_path = 'model_present_death.pkl'
model_progression_path = 'model_present_progression.pkl'

if os.path.exists(model_death_path):
    model_death = load(model_death_path)
    print("Model 'model_present_death.pkl' carregat correctament.")
else:
    raise FileNotFoundError(f"El fitxer del model '{model_death_path}' no s'ha trobat.")

if os.path.exists(model_progression_path):
    model_progression = load(model_progression_path)
    print("Model 'model_present_progression.pkl' carregat correctament.")
else:
    raise FileNotFoundError(f"El fitxer del model '{model_progression_path}' no s'ha trobat.")

# Variables esperades pel model (assegura't que aquestes coincideixen amb les del model entrenat)
model_variables = [
                'Usuari', 'Pedigree', 'Sex', 'Age at diagnosis', 'Final diagnosis', 
                'TOBACCO', 'Radiological Pattern', 'Biopsy', 'Extrapulmonary affectation',
                'Associated lung cancer', 'Other cancer', 'Type of neoplasia',
                'Hematological abnormality before diagnosis','Blood count abnormality at diagnosis', 
                'Anemia', 'Thrombocytopenia', 'Thrombocytosis', 'Lymphocytosis',
                'Lymphopenia', 'Neutrophilia', 'Neutropenia', 'Leukocytosis', 'Leukopenia',
                'Hematologic Disease', 'Liver abnormality before diagnosis', 'Liver abnormality',
                'LDH', 'ALT', 'AST', 'ALP', 'GGT', 'Transaminitis', 'Cholestasis', 'Liver disease',
                'FVC (%) at diagnosis', 'DLCO (%) at diagnosis', '1st degree relative', '2nd degree relative',
                'More than 1 relative', 'Genetic mutation studied in patient', 
                'Severity of telomere shortening'
            ]

# Definir mapejos per a les variables categòriques
pedigree_mapping = {
    '1': 1,  # Suposant '1' és 'Familial'
    '0': 0   # Suposant '0' és 'Sporadic'
}

sex_mapping = {
    'Male': 1,
    'Female': 0
}

final_diagnosis_mapping = {
    'No IPF': 0,
    'IPF': 1,
    'CHP': 2,
    'SRIF': 3,
    'NSIP': 4,
    'CPFE': 5,
    'PF-CTD (RA)': 6,
    'PF-CTD (SLE)': 7,
    'Incipient': 8,
    'Other': 9  # Afegir altres diagnoses si escau
}

radiological_pattern_mapping = {
    'Non UIP': 0,
    'UIP': 1,
    'Indeterminate UIP': 2,
    'Probable UIP': 3,
    'Unknown': 4  # Afegir altres patterns si escau
}

tobacco_mapping = {
    'No history of smoking': 0,
    'Active smoker': 1,
    'Ex-smoker': 2,
    'Unknown': -1  # Assignar -1 per a casos no reconeguts
}

biopsy_mapping = {
    'biopsy-none': 0,
    'biopsy-endoscopic': 1,
    'biopsy-surgical': 2,
    'Unknown': -1  # Assignar -1 per a casos no reconeguts
}

extrapulmonary_mapping = {
    'no': 0,
    'yes': 1
}

cancer_mapping = {
    'no': 0,
    'yes': 1
}

hematologic_abnormalities_mapping = {
    'no': 0,
    'yes': 1
}

hematologic_disease_confirm_mapping = {
    'no': 0,
    'yes': 1
}

liver_abnormality_before_mapping = {
    'no': 0,
    'yes': 1
}

liver_abnormality_mapping = {
    'no': 0,
    'yes': 1
}

liver_disease_mapping = {
    'no': 0,
    'yes': 1
}

telomere_shortening_mapping = {
    'none': 0,
    'mild': 1,
    'moderate': 2,
    'severe': 3,
    'unknown': -1  # Assignar -1 per a casos no reconeguts
}

# Ruta per a la pàgina principal
@app.route('/')
def index():
    return render_template('app_prova.html')  # Assegura't que aquest fitxer HTML està a /templates

# Ruta per guardar les dades al fitxer Excel
@app.route('/save-excel', methods=['POST'])
def save_excel():
    try:
        data = request.get_json()  # Rebre dades en format JSON
        print(f"Dades rebudes: {data}")  # Per depuració

        if not data:
            return jsonify({'error': 'No hi ha dades per guardar.'}), 400

        # Convertir les llistes a strings, o usar 'None' si estan buides
        data['NeoplasiaType'] = ', '.join(data.get('NeoplasiaType', ['None']))
        data['BloodCountAbnormalities'] = ', '.join(data.get('BloodCountAbnormalities', ['None']))
        data['HematologicDiseaseTypes'] = ', '.join(data.get('HematologicDiseaseTypes', ['None']))
        data['MutationType'] = ', '.join(data.get('MutationType', ['None']))

        # Codificar les dades de l'usuari
        encoded_data = {}

        # Codificar 'Pedigree'
        pedigree = data.get('Pedigree', '0')  # Suposant '0' és 'Sporadic' per defecte
        encoded_data['Pedigree'] = pedigree_mapping.get(pedigree, -1)

        # Codificar 'sex'
        sex = data.get('sex', 'Male')  # Suposant 'Male' per defecte
        encoded_data['sex'] = sex_mapping.get(sex, -1)

        # 'Age at diagnosis' assumeixo que és numèric
        age = data.get('Age at diagnosis', None)
        if age is not None:
            try:
                encoded_data['Age at diagnosis'] = float(age)
            except ValueError:
                encoded_data['Age at diagnosis'] = -1  # Assignar -1 per a dades no vàlides
        else:
            encoded_data['Age at diagnosis'] = -1  # Assignar -1 per a dades faltants

        # Codificar 'FinalDiagnosis'
        final_diagnosis = data.get('FinalDiagnosis', 'Other')
        encoded_data['FinalDiagnosis'] = final_diagnosis_mapping.get(final_diagnosis, 9)

        # Codificar 'TobaccoHistory'
        tobacco_history = data.get('TobaccoHistory', 'Unknown')
        encoded_data['TobaccoHistory'] = tobacco_mapping.get(tobacco_history, -1)

        # Codificar 'RadiologicalPattern'
        radiological_pattern = data.get('RadiologicalPattern', 'Unknown')
        encoded_data['RadiologicalPattern'] = radiological_pattern_mapping.get(radiological_pattern, 4)

        # Codificar 'Biopsy'
        biopsy = data.get('Biopsy', 'Unknown')
        encoded_data['Biopsy'] = biopsy_mapping.get(biopsy, -1)

        # Codificar 'Extrapulmonary'
        extrapulmonary = data.get('Extrapulmonary', 'no')
        encoded_data['Extrapulmonary'] = extrapulmonary_mapping.get(extrapulmonary.lower(), -1)

        # Codificar 'LungCancer'
        lung_cancer = data.get('LungCancer', 'no')
        encoded_data['LungCancer'] = cancer_mapping.get(lung_cancer.lower(), -1)

        # Codificar 'OtherCancer'
        other_cancer = data.get('OtherCancer', 'no')
        encoded_data['OtherCancer'] = cancer_mapping.get(other_cancer.lower(), -1)

        # Codificar 'NeoplasiaType'
        encoded_data['NeoplasiaType'] = data.get('NeoplasiaType', 'None')

        # Codificar 'HematologicAbnormalities'
        hematologic_abnormalities = data.get('HematologicAbnormalities', 'no')
        encoded_data['HematologicAbnormalities'] = hematologic_abnormalities_mapping.get(hematologic_abnormalities.lower(), -1)

        # Codificar 'BloodCountAbnormalities'
        encoded_data['BloodCountAbnormalities'] = data.get('BloodCountAbnormalities', 'None')

        # Codificar 'HematologicDiseaseConfirm'
        hematologic_disease_confirm = data.get('HematologicDiseaseConfirm', 'no')
        encoded_data['HematologicDiseaseConfirm'] = hematologic_disease_confirm_mapping.get(hematologic_disease_confirm.lower(), -1)

        # Codificar 'HematologicDiseaseTypes'
        encoded_data['HematologicDiseaseTypes'] = data.get('HematologicDiseaseTypes', 'None')

        # Codificar 'LiverAbnormalityBefore'
        liver_abnormality_before = data.get('LiverAbnormalityBefore', 'no')
        encoded_data['LiverAbnormalityBefore'] = liver_abnormality_before_mapping.get(liver_abnormality_before.lower(), -1)

        # Codificar 'LiverAbnormality'
        liver_abnormality = data.get('LiverAbnormality', 'no')
        encoded_data['LiverAbnormality'] = liver_abnormality_mapping.get(liver_abnormality.lower(), -1)

        # Codificar 'LiverDisease'
        liver_disease = data.get('LiverDisease', 'no')
        encoded_data['LiverDisease'] = liver_disease_mapping.get(liver_disease.lower(), -1)

        # Codificar 'LDH', 'ALT', 'AST', 'ALP', 'GGT', 'Transaminitis', 'Cholestasis'
        # Aquí assumirem que 'normal' = 0, i altres valors poden ser mapejats si escau
        lab_mapping = {
            'normal': 0,
            'abnormal': 1,
            # Afegir altres valors si escau
        }
        lab_variables = ['LDH', 'ALT', 'AST', 'ALP', 'GGT', 'Transaminitis', 'Cholestasis']
        for lab in lab_variables:
            value = data.get(lab, 'normal').lower()
            encoded_data[lab] = lab_mapping.get(value, -1)

        # 'FVC' i 'DLCO' són numèrics
        try:
            encoded_data['FVC'] = float(data.get('FVC', '0'))
        except ValueError:
            encoded_data['FVC'] = -1

        try:
            encoded_data['DLCO'] = float(data.get('DLCO', '0'))
        except ValueError:
            encoded_data['DLCO'] = -1

        # Codificar 'FirstDegreeRelative', 'SecondDegreeRelative', 'MoreThanOneRelative'
        relative_mapping = {
            'no': 0,
            'yes': 1
        }
        encoded_data['FirstDegreeRelative'] = relative_mapping.get(data.get('FirstDegreeRelative', 'no').lower(), -1)
        encoded_data['SecondDegreeRelative'] = relative_mapping.get(data.get('SecondDegreeRelative', 'no').lower(), -1)
        encoded_data['MoreThanOneRelative'] = relative_mapping.get(data.get('MoreThanOneRelative', 'no').lower(), -1)

        # Codificar 'GeneticMutation'
        genetic_mutation = data.get('GeneticMutation', 'no')
        encoded_data['GeneticMutation'] = relative_mapping.get(genetic_mutation.lower(), -1)

        # Codificar 'TelomereShorteningSeverity'
        telomere_shortening = data.get('TelomereShorteningSeverity', 'none').lower()
        encoded_data['TelomereShorteningSeverity'] = telomere_shortening_mapping.get(telomere_shortening, -1)

        print(f"Dades codificades: {encoded_data}")  # Per depuració

        # Crear o carregar el fitxer Excel
        filepath = os.path.join(SAVE_FOLDER, 'respostes_questionari.xlsx')
        if os.path.exists(filepath):
            workbook = load_workbook(filepath)
        else:
            workbook = Workbook()

        # Obtenir o crear la fulla de respostes
        sheet_name = 'Respostes'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
        
        # Afegir els encapçalaments si no existeixen
        if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
            headers = [
                'Usuari', 'Pedigree', 'Sex', 'Age at diagnosis', 'Final diagnosis', 
                'TOBACCO', 'Radiological Pattern', 'Biopsy', 'Extrapulmonary affectation',
                'Associated lung cancer', 'Other cancer', 'Type of neoplasia',
                'Hematological abnormality before diagnosis','Blood count abnormality at diagnosis', 
                'Anemia', 'Thrombocytopenia', 'Thrombocytosis', 'Lymphocytosis',
                'Lymphopenia', 'Neutrophilia', 'Neutropenia', 'Leukocytosis', 'Leukopenia',
                'Hematologic Disease', 'Liver abnormality before diagnosis', 'Liver abnormality',
                'LDH', 'ALT', 'AST', 'ALP', 'GGT', 'Transaminitis', 'Cholestasis', 'Liver disease',
                'FVC (%) at diagnosis', 'DLCO (%) at diagnosis', '1st degree relative', '2nd degree relative',
                'More than 1 relative', 'Genetic mutation studied in patient', 
                'Severity of telomere shortening'
            ]
     
            sheet.append(headers)

        # Afegir la nova fila amb les dades
        user = data.get('usuari', 'Desconegut')
        # Crear la variable 'Blood count abnormality at diagnosis'
        # Codificar 'BloodCountAbnormalities'
        blood_count_options = [
            'anemia', 'thrombocytopenia', 'thrombocytosis', 'lymphocytosis',
            'lymphopenia', 'neutrophilia', 'neutropenia', 'leukocytosis', 'leukopenia'
        ]
        selected_blood_abnormalities = data.get('BloodCountAbnormalities', [])

        # Crear una columna para cada opción con valores binarios
        for abnormality in blood_count_options:
            encoded_data[abnormality] = 1 if abnormality in selected_blood_abnormalities else 0

        encoded_data['BloodCountAbnormalityAtDiagnosis'] = int(
            any([
                encoded_data['anemia'], 
                encoded_data['thrombocytopenia'], 
                encoded_data['thrombocytosis'], 
                encoded_data['lymphocytosis'], 
                encoded_data['lymphopenia'], 
                encoded_data['neutrophilia'], 
                encoded_data['neutropenia'], 
                encoded_data['leukocytosis'], 
                encoded_data['leukopenia']
            ])
        )

        row = [
            user,
            encoded_data['Pedigree'], encoded_data['sex'], encoded_data['Age at diagnosis'],
            encoded_data['FinalDiagnosis'], encoded_data['TobaccoHistory'],
            encoded_data['RadiologicalPattern'], encoded_data['Biopsy'], encoded_data['Extrapulmonary'],
            encoded_data['LungCancer'], encoded_data['OtherCancer'], encoded_data['NeoplasiaType'],
            encoded_data['HematologicAbnormalities'],
            encoded_data['BloodCountAbnormalityAtDiagnosis'],
            # Valores binarios para cada anomalía sanguínea
            encoded_data['anemia'], encoded_data['thrombocytopenia'], encoded_data['thrombocytosis'],
            encoded_data['lymphocytosis'], encoded_data['lymphopenia'], encoded_data['neutrophilia'],
            encoded_data['neutropenia'], encoded_data['leukocytosis'], encoded_data['leukopenia'],
            # Otros valores
            encoded_data['HematologicDisease'], encoded_data['LiverAbnormalityBefore'], encoded_data['LiverAbnormality'],
            encoded_data['LDH'], encoded_data['ALT'], encoded_data['AST'], encoded_data['ALP'],
            encoded_data['GGT'], encoded_data['Transaminitis'], encoded_data['Cholestasis'], encoded_data['LiverDisease'],
            encoded_data['FVC'], encoded_data['DLCO'], encoded_data['FirstDegreeRelative'], 
            encoded_data['SecondDegreeRelative'], encoded_data['MoreThanOneRelative'],
            encoded_data['GeneticMutation'], encoded_data['TelomereShorteningSeverity']
        ]

        sheet.append(row)

        # Guardar l'Excel
        workbook.save(filepath)

        # Crear DataFrame per a la predicció
        df_input = pd.DataFrame([encoded_data], columns=model_variables)


        # Ensure all columns expected by the models are present
        # Per model_death
        try:
            model_death_columns = model_death.feature_names_in_
            print(f"Columnes esperades pel model_death: {model_death_columns}")
        except AttributeError:
            # Si el model no té 'feature_names_in_', usar les columnes del DataFrame
            model_death_columns = df_input.columns
            print("Model_death no té 'feature_names_in_', s'utilitzen les columnes del DataFrame.")

          # Per model_progression
        try:
            model_progression_columns = model_progression.feature_names_in_
            print(f"Columnes esperades pel model_progression: {model_progression_columns}")
        except AttributeError:
            # Si el model no té 'feature_names_in_', usar les columnes del DataFrame
            model_progression_columns = df_input.columns
            print("Model_progression no té 'feature_names_in_', s'utilitzen les columnes del DataFrame.")

        # Afegir columnes mancants amb valor 0 per al model_death
        missing_cols_death = set(model_death_columns) - set(df_input.columns)
        if missing_cols_death:
            for col in missing_cols_death:
                df_input[col] = 0
            print(f"Columnes mancants afegides amb valor 0 per a model_death: {missing_cols_death}")

        # Afegir columnes mancants amb valor 0 per al model_progression
        missing_cols_progression = set(model_progression_columns) - set(df_input.columns)
        if missing_cols_progression:
            for col in missing_cols_progression:
                df_input[col] = 0
            print(f"Columnes mancants afegides amb valor 0 per a model_progression: {missing_cols_progression}")

        # Reordenar les columnes per coincidir amb els models
        df_death_input = df_input.reindex(columns=model_death_columns, fill_value=0)
        df_progression_input = df_input.reindex(columns=model_progression_columns, fill_value=0)
        print(f"DataFrame reordenat per a model_death:\n{df_death_input}")
        print(f"DataFrame reordenat per a model_progression:\n{df_progression_input}")

        # Fer les prediccions
        prediction_death = model_death.predict(df_death_input)[0]
        prediction_progression = model_progression.predict(df_progression_input)[0]
        print(f"Predicció model_death: {prediction_death}")
        print(f"Predicció model_progression: {prediction_progression}")

        # Afegir les prediccions a la fila corresponent al fitxer Excel
        # Suposant que les prediccions es guarden en les dues últimes columnes
        last_row = sheet.max_row
        sheet.cell(row=last_row, column=len(row) + 1).value = prediction_death
        sheet.cell(row=last_row, column=len(row) + 2).value = prediction_progression
        print("Prediccions afegides al fitxer Excel.")

        # Guardar l'Excel amb les prediccions
        workbook.save(filepath)
        print("Fitxer Excel guardat amb les prediccions.")

        return jsonify({
            'message': 'Dades guardades correctament i prediccions fetes.', 
            'predicció_death': int(prediction_death),
            'predicció_progression': int(prediction_progression)
        }), 200

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': 'Hi ha hagut un problema en desar les dades o fer les prediccions.'}), 500
    
if __name__ == '__main__':
    app.run(debug=True)