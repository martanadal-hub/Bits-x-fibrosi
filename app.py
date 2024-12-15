from flask import Flask, request, jsonify, render_template
from openpyxl import Workbook, load_workbook
from joblib import load
import pandas as pd
import os

app = Flask(__name__)

# Crear la carpeta de dades
SAVE_FOLDER = 'BITS-X-FIBROSI'
if not os.path.exists(SAVE_FOLDER):
    os.makedirs(SAVE_FOLDER)

# Carregar el model entrenat
model_path = 'model_death.pkl'
if os.path.exists(model_path):
    model = load(model_path)
    print("Model carregat correctament.")
else:
    raise FileNotFoundError(f"El fitxer del model '{model_path}' no s'ha trobat.")

# Variables esperades pel model (assegura't que aquestes coincideixen amb les del model entrenat)
model_variables = [
    'Pedigree', 'sex', 'Age at diagnosis', 'FinalDiagnosis', 
    'TobaccoHistory', 'RadiologicalPattern', 'Biopsy',
    'Extrapulmonary', 'LungCancer', 'OtherCancer', 'NeoplasiaType',
    'HematologicAbnormalities', 'BloodCountAbnormalities', 
    'HematologicDiseaseConfirm', 'HematologicDiseaseTypes', 
    'LiverAbnormalityBefore', 'LiverAbnormality', 'LiverDisease',
    'LDH', 'ALT', 'AST', 'ALP', 'GGT', 'Transaminitis', 'Cholestasis',
    'FVC', 'DLCO', 'FirstDegreeRelative', 'SecondDegreeRelative',
    'MoreThanOneRelative', 'GeneticMutation',  
    'TelomereShorteningSeverity'
]

# Definir mapejos per a les variables categòriques
sex_mapping = {
    'Dona': 'Male',
    'Home': 'Female'
}

tobacco_mapping = {
    'Sense antecedents de tabaquisme': 0,
    'Fumador actiu': 1,
    'Ex-fumat': 2
}

radiological_pattern_mapping = {
    'Non UIP': 'Non UIP',
    'UIP': 'UIP',
    'Indeterminate UIP': 'Indeterminate UIP',
    'Probable UIP': 'Probable UIP'
}

biopsy_mapping = {
    'Sense biopsia': 0,
    'Criobiòpsia endoscòpica': 1,
    'Biopsia quirúrgica': 2
}

extrapulmonary_mapping = {
    'no': 0,
    'yes': 1
}

cancer_mapping = {
    'no': 0,
    'yes': 1
}

hematologic_abnormalities_mapping = {
    'no': 0,
    'yes': 1
}

hematologic_disease_confirm_mapping = {
    'no': 0,
    'yes': 1
}

liver_abnormality_before_mapping = {
    'no': 0,
    'yes': 1
}

liver_abnormality_mapping = {
    'no': 0,
    'yes': 1
}

liver_disease_mapping = {
    'no': 0,
    'yes': 1
}

final_diagnosis_mapping = {
    'No IPF': 0,
    'IPF': 1,
    'CHP': 2,
    'SRIF': 3,
    'NSIP': 4,
    'CPFE': 5,
    'PF-CTD (RA)': 6,
    'PF-CTD (SLE)': 7,
    'Incipient': 8,
    'Other': 9  # Afegir altres diagnoses si escau
}


# Ruta per a la pàgina principal
@app.route('/')
def index():
    return render_template('app_prova.html')  # Assegura't que aquest fitxer HTML està a /templates

# Ruta per guardar les dades al fitxer Excel
@app.route('/save-excel', methods=['POST'])
def save_excel():
    try:
        data = request.get_json()  # Rebre dades en format JSON
        print(f"Dades rebudes: {data}")  # Per depuració

        if not data:
            return jsonify({'error': 'No hi ha dades per guardar.'}), 400

        # Convertir les llistes a strings, o usar 'None' si estan buides
        data['NeoplasiaType'] = ', '.join(data.get('NeoplasiaType', ['None']))
        data['BloodCountAbnormalities'] = ', '.join(data.get('BloodCountAbnormalities', ['None']))
        data['HematologicDiseaseTypes'] = ', '.join(data.get('HematologicDiseaseTypes', ['None']))
        data['MutationType'] = ', '.join(data.get('MutationType', ['None']))

        # Codificar les dades de l'usuari
        encoded_data = {}

        # Codificar 'sex'
        sex = data.get('sex', 'Male')  # Suposant 'Male' per defecte
        encoded_data['sex'] = sex_mapping.get(sex, -1)

        # 'Age at diagnosis' assumeixo que és numèric
        age = data.get('Age at diagnosis', None)
        if age is not None:
            try:
                encoded_data['Age at diagnosis'] = float(age)
            except ValueError:
                encoded_data['Age at diagnosis'] = -1  # Assignar -1 per a dades no vàlides
        else:
            encoded_data['Age at diagnosis'] = -1  # Assignar -1 per a dades faltants

        # Codificar 'FinalDiagnosis'
        final_diagnosis = data.get('FinalDiagnosis', 'Other')
        encoded_data['FinalDiagnosis'] = final_diagnosis_mapping.get(final_diagnosis, 9)

        # Codificar 'TobaccoHistory'
        tobacco_history = data.get('TobaccoHistory', 'Unknown')
        encoded_data['TobaccoHistory'] = tobacco_mapping.get(tobacco_history, -1)

        # Codificar 'RadiologicalPattern'
        radiological_pattern = data.get('RadiologicalPattern', 'Unknown')
        encoded_data['RadiologicalPattern'] = radiological_pattern_mapping.get(radiological_pattern, 4)

        # Codificar 'Biopsy'
        biopsy = data.get('Biopsy', 'Unknown')
        encoded_data['Biopsy'] = biopsy_mapping.get(biopsy, -1)

        # Codificar 'Extrapulmonary'
        extrapulmonary = data.get('Extrapulmonary', 'no')
        encoded_data['Extrapulmonary'] = extrapulmonary_mapping.get(extrapulmonary.lower(), -1)

        # Codificar 'LungCancer'
        lung_cancer = data.get('LungCancer', 'no')
        encoded_data['LungCancer'] = cancer_mapping.get(lung_cancer.lower(), -1)

        # Codificar 'OtherCancer'
        other_cancer = data.get('OtherCancer', 'no')
        encoded_data['OtherCancer'] = cancer_mapping.get(other_cancer.lower(), -1)

        # Codificar 'NeoplasiaType'
        encoded_data['NeoplasiaType'] = data.get('NeoplasiaType', 'None')

        # Codificar 'HematologicAbnormalities'
        hematologic_abnormalities = data.get('HematologicAbnormalities', 'no')
        encoded_data['HematologicAbnormalities'] = hematologic_abnormalities_mapping.get(hematologic_abnormalities.lower(), -1)

        # Codificar 'BloodCountAbnormalities'
        encoded_data['BloodCountAbnormalities'] = data.get('BloodCountAbnormalities', 'None')

        # Codificar 'HematologicDiseaseConfirm'
        hematologic_disease_confirm = data.get('HematologicDiseaseConfirm', 'no')
        encoded_data['HematologicDiseaseConfirm'] = hematologic_disease_confirm_mapping.get(hematologic_disease_confirm.lower(), -1)

        # Codificar 'HematologicDiseaseTypes'
        encoded_data['HematologicDiseaseTypes'] = data.get('HematologicDiseaseTypes', 'None')

        # Codificar 'LiverAbnormalityBefore'
        liver_abnormality_before = data.get('LiverAbnormalityBefore', 'no')
        encoded_data['LiverAbnormalityBefore'] = liver_abnormality_before_mapping.get(liver_abnormality_before.lower(), -1)

        # Codificar 'LiverAbnormality'
        liver_abnormality = data.get('LiverAbnormality', 'no')
        encoded_data['LiverAbnormality'] = liver_abnormality_mapping.get(liver_abnormality.lower(), -1)

        # Codificar 'LiverDisease'
        liver_disease = data.get('LiverDisease', 'no')
        encoded_data['LiverDisease'] = liver_disease_mapping.get(liver_disease.lower(), -1)

        # Codificar 'LDH', 'ALT', 'AST', 'ALP', 'GGT', 'Transaminitis', 'Cholestasis'
        # Aquí assumirem que 'normal' = 0, i altres valors poden ser mapejats si escau
        lab_mapping = {
            'normal': 0,
            'abnormal': 1,
            # Afegir altres valors si escau
        }
        lab_variables = ['LDH', 'ALT', 'AST', 'ALP', 'GGT', 'Transaminitis', 'Cholestasis']
        for lab in lab_variables:
            value = data.get(lab, 'normal').lower()
            encoded_data[lab] = lab_mapping.get(value, -1)

        # 'FVC' i 'DLCO' són numèrics
        try:
            encoded_data['FVC'] = float(data.get('FVC', '0'))
        except ValueError:
            encoded_data['FVC'] = -1

        try:
            encoded_data['DLCO'] = float(data.get('DLCO', '0'))
        except ValueError:
            encoded_data['DLCO'] = -1

        # Codificar 'FirstDegreeRelative', 'SecondDegreeRelative', 'MoreThanOneRelative'
        relative_mapping = {
            'No': 0,
            'Sí': 1
        }
        encoded_data['FirstDegreeRelative'] = relative_mapping.get(data.get('FirstDegreeRelative', 'no').lower(), -1)
        encoded_data['SecondDegreeRelative'] = relative_mapping.get(data.get('SecondDegreeRelative', 'no').lower(), -1)
        encoded_data['MoreThanOneRelative'] = relative_mapping.get(data.get('MoreThanOneRelative', 'no').lower(), -1)

        # Codificar 'GeneticMutation'
        genetic_mutation = data.get('GeneticMutation', 'no')
        encoded_data['GeneticMutation'] = relative_mapping.get(genetic_mutation.lower(), -1)

        print(f"Dades codificades: {encoded_data}")  # Per depuració

        # Crear o carregar el fitxer Excel
        filepath = os.path.join(SAVE_FOLDER, 'respostes_questionari.xlsx')
        if os.path.exists(filepath):
            workbook = load_workbook(filepath)
        else:
            workbook = Workbook()

        # Obtenir o crear la fulla de respostes
        sheet_name = 'Respostes'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)

        # Afegir els encapçalaments si no existeixen
        if sheet.max_row == 1 and all(cell.value is None for cell in sheet[1]):
            headers = [
                'Usuari', 'Pedigree', 'sex', 'Age at diagnosis', 'FinalDiagnosis', 
                'TobaccoHistory', 'RadiologicalPattern', 'Biopsy',
                'Extrapulmonary', 'LungCancer', 'OtherCancer', 'NeoplasiaType',
                'HematologicAbnormalities', 'BloodCountAbnormalities', 
                'HematologicDiseaseConfirm', 'HematologicDiseaseTypes', 
                'LiverAbnormalityBefore', 'LiverAbnormality', 'LiverDisease',
                'LDH', 'ALT', 'AST', 'ALP', 'GGT', 'Transaminitis', 'Cholestasis',
                'FVC', 'DLCO', 'FirstDegreeRelative', 'SecondDegreeRelative',
                'MoreThanOneRelative', 'GeneticMutation', 
                'TelomereShorteningSeverity', 'Predicció'
            ]
            sheet.append(headers)

        # Afegir la nova fila amb les dades
        user = data.get('usuari', 'Desconegut')
        row = [
            user,
            encoded_data['Pedigree'],
            encoded_data['sex'],
            encoded_data['Age at diagnosis'],
            encoded_data['FinalDiagnosis'],
            encoded_data['TobaccoHistory'],
            encoded_data['RadiologicalPattern'],
            encoded_data['Biopsy'],
            encoded_data['Extrapulmonary'],
            encoded_data['LungCancer'],
            encoded_data['OtherCancer'],
            encoded_data['NeoplasiaType'],
            encoded_data['HematologicAbnormalities'],
            encoded_data['BloodCountAbnormalities'],
            encoded_data['HematologicDiseaseConfirm'],
            encoded_data['HematologicDiseaseTypes'],
            encoded_data['LiverAbnormalityBefore'],
            encoded_data['LiverAbnormality'],
            encoded_data['LiverDisease'],
            encoded_data['LDH'],
            encoded_data['ALT'],
            encoded_data['AST'],
            encoded_data['ALP'],
            encoded_data['GGT'],
            encoded_data['Transaminitis'],
            encoded_data['Cholestasis'],
            encoded_data['FVC'],
            encoded_data['DLCO'],
            encoded_data['FirstDegreeRelative'],
            encoded_data['SecondDegreeRelative'],
            encoded_data['MoreThanOneRelative'],
            encoded_data['GeneticMutation'],
            encoded_data['TelomereShorteningSeverity']
        ]
        sheet.append(row)

        # Guardar l'Excel
        workbook.save(filepath)

        # Crear DataFrame per a la predicció
        df_input = pd.DataFrame([encoded_data], columns=model_variables)

        # Aplicar One-Hot Encoding per a variables com 'NeoplasiaType', 'BloodCountAbnormalities', 'HematologicDiseaseTypes', 'MutationType'
        # Convertir-les a variables binàries (multi-label)
        categorical_vars = ['NeoplasiaType', 'BloodCountAbnormalities', 'HematologicDiseaseTypes']
        for var in categorical_vars:
            dummies = df_input[var].str.get_dummies(sep=', ')
            df_input = pd.concat([df_input, dummies], axis=1)
            df_input.drop(var, axis=1, inplace=True)

        # Ensure all columns expected by the model are present
        try:
            # Intentar obtenir les columnes del model
            model_columns = model.feature_names_in_
        except AttributeError:
            # Si el model no té 'feature_names_in_', usar les columnes del DataFrame
            model_columns = df_input.columns

        # Afegir columnes mancants amb valor 0
        missing_cols = set(model_columns) - set(df_input.columns)
        for col in missing_cols:
            df_input[col] = 0

        # Reordenar les columnes per coincidir amb el model
        df_input = df_input[model_columns]

        # Preprocessar les dades noves per predicció
        prediction = model.predict(df_input)[0]
        print(f"Predicció: {prediction}")  # Per depuració

        # Afegir la predicció a la fila corresponent al fitxer Excel
        sheet.cell(row=sheet.max_row, column=len(row) + 1).value = prediction
        workbook.save(filepath)

        return jsonify({'message': 'Dades guardades correctament i predicció feta.', 'predicció': int(prediction)}), 200

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': 'Hi ha hagut un problema en desar les dades o fer la predicció.'}), 500

if __name__ == '__main__':
    app.run(debug=True)
